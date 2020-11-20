import os
import openpyxl
from typing import Union

from HSTB.kluster.fqpr_generation import Fqpr
from HSTB.kluster.fqpr_convenience import process_and_export_soundings, generate_new_surface, reload_data
from HSTB.kluster.fqpr_surface import BaseSurface

mode_translator = {'vsCW': 'CW_veryshort', 'shCW': 'CW_short', 'meCW': 'CW_medium', 'loCW': 'CW_long',
                   'vlCW': 'CW_verylong', 'elCW': 'CW_extralong', 'shFM': 'FM_short', 'loFM': 'FM_long',
                   '__FM': 'FM', 'FM': 'FM', 'CW': 'CW', 'VS': 'VeryShallow', 'SH': 'Shallow', 'ME': 'Medium',
                   'DE': 'Deep', 'VD': 'VeryDeep', 'ED': 'ExtraDeep'}


def return_files_from_path(pth: str, file_ext: str = '.all'):
    """
    Input files can be entered into an xarray_conversion.BatchRead instance as either a list, a path to a directory
    of multibeam files or as a path to a single file.  Here we return all the files in each of these scenarios as a list
    for the gui to display or to be analyzed in some other way

    Provide an optional fileext argument if you want to specify files with a different extension

    Parameters
    ----------
    pth
        either a list of files, a string path to a directory or a string path to a file
    file_ext
        file extension of the file(s) you are looking for

    Returns
    -------
    list
        list of files found
    """

    if type(pth) == list:
        if len(pth) == 1 and os.path.isdir(pth[0]):  # a list one element long that is a path to a directory
            return [os.path.join(pth[0], p) for p in os.listdir(pth[0]) if os.path.splitext(p)[1] == file_ext]
        else:
            return [p for p in pth if os.path.splitext(p)[1] == file_ext]
    elif os.path.isdir(pth):
        return [os.path.join(pth, p) for p in os.listdir(pth) if os.path.splitext(p)[1] == file_ext]
    elif os.path.isfile(pth):
        if os.path.splitext(pth)[1] == file_ext:
            return [pth]
        else:
            return []
    else:
        return []


def return_directory_from_data(data: Union[list, str]):
    """
    Given either a path to a zarr store, a path to a directory of multibeam files, a list of paths to multibeam files
    or a path to a single multibeam file, return the parent directory.

    Parameters
    ----------
    data
        a path to a zarr store, a path to a directory of multibeam files, a list of paths to multibeam files or a path
        to a single multibeam file.

    Returns
    -------
    output_directory: str, path to output directory
    """

    try:  # they provided a path to a zarr store or a path to a directory of .all files or a single .all file
        output_directory = os.path.dirname(data)
    except TypeError:  # they provided a list of files
        output_directory = os.path.dirname(data[0])
    return output_directory


def return_data(pth: Union[list, str], coord_system: str = 'NAD83', vert_ref: str = 'waterline',
                require_raw_data: bool = True, autogenerate: bool = True, skip_dask: bool = False):
    """
    Take in a path to a zarr store, a path to a directory of multibeam files, a list of paths to multibeam files or a
    path to a single multibeam file and return a loaded or newly constructed fqpr_generation.Fqpr instance.

    Parameters
    ----------
    pth
        a path to a zarr store, a path to a directory of .all files, a list of paths to .all files or
          a path to a single .all file.
    coord_system
        one of ['NAD83', 'WGS84']
    vert_ref
        one of ['waterline', 'ellipse', 'vessel']
    require_raw_data
        if True, raise exception if you can't find the raw data
    autogenerate
        if True will build a new xyz dataset if a path is passed in
    skip_dask
        if True, will not start/find the dask client.  Only use this if you are just reading attribution

    Returns
    -------
    fqpr_generation.Fqpr
        processed or loaded instance for the given data
    """

    fqpr_instance = None
    try:
        fqpr_instance = reload_data(pth, require_raw_data=require_raw_data, skip_dask=skip_dask)
    except (TypeError, ValueError):
        if autogenerate:
            fqpr_instance = process_and_export_soundings(pth, coord_system=coord_system, vert_ref=vert_ref)
    return fqpr_instance


def return_surface(ref_surf_pth: Union[list, str], vert_ref: str, resolution: int, autogenerate: bool = True):
    """
    Take in a path to a zarr store, a path to a directory of multibeam files, a list of paths to multibeam files, a path
    to a single multibeam file or a path to an existing surface.  Return a loaded or newly constructed fqpr_surface
    BaseSurface instance.

    Parameters
    ----------
    ref_surf_pth
        a path to a zarr store, a path to a directory of multibeam files, a list of paths to multibeam files, a path to
        a single multibeam file or a path to an existing surface.
    vert_ref
        one of ['waterline', 'ellipse', 'vessel']
    resolution
        resolution of the grid in meters
    autogenerate
        if True will build a new surface if a path is passed in

    Returns
    -------
    fqpr_surface.BaseSurface
        surface for the given data
    """

    need_surface = False
    bs = None

    if os.path.isfile(ref_surf_pth):
        if os.path.splitext(ref_surf_pth)[1] not in ['.all', '.kmall']:
            bs = BaseSurface(from_file=ref_surf_pth)
        else:
            need_surface = True
    else:
        need_surface = True

    if need_surface and autogenerate:
        new_surf_path = os.path.join(return_directory_from_data(ref_surf_pth), 'surface_{}m.npy'.format(resolution))
        fqpr_instance = return_data(ref_surf_pth, vert_ref, autogenerate=True)
        bs = generate_new_surface(fqpr_instance, resolution, client=fqpr_instance.client)
        bs.save(new_surf_path)
    return bs


def get_attributes_from_zarr_stores(list_dir_paths: list):
    """
    Takes in a list of paths to directories containing fqpr generated zarr stores.  Returns a list where each element
    is a dict of attributes found in each zarr store.  Prefers the attributes from the xyz_dat store, but if it can't
    find it, will read from one of the raw_ping converted zarr stores.  (all attributes across raw_ping data stores
    are identical)

    Parameters
    ----------
    list_dir_paths
        list of strings for paths to each converted folder containing the zarr folders

    Returns
    -------
    list
        list of dicts for each successfully reloaded fqpr object
    """

    attrs = []
    for pth in list_dir_paths:
        fqpr_instance = reload_data(pth, skip_dask=True)
        if fqpr_instance is not None:
            newattrs = get_attributes_from_fqpr(fqpr_instance)
            attrs.append(newattrs)
        else:
            attrs.append([None])
    return attrs


def get_attributes_from_fqpr(fqpr_instance: Fqpr, include_mode: bool = True):
    """
    Takes in a FQPR instance.  Returns a dict of the attribution in that instance.  Prefers the attributes from the
    xyz_dat store, but if it can't find it, will read from one of the raw_ping converted zarr stores.
    (all attributes across raw_ping data stores are identical)

    If include_mode is True, will also read and translate unique modes from each raw_ping and include a unique set of
    those modes in the returned attributes.

    Parameters
    ----------
    fqpr_instance
        fqpr instance that you want to load from
    include_mode
        if True, include mode in the returned attributes

    Returns
    -------
    dict
        dict of attributes in that FQPR instance
    """

    if 'xyz_dat' in fqpr_instance.__dict__:
        if fqpr_instance.soundings is not None:
            newattrs = fqpr_instance.soundings.attrs
        else:
            newattrs = fqpr_instance.multibeam.raw_ping[0].attrs
    else:
        newattrs = fqpr_instance.multibeam.raw_ping[0].attrs

    try:
        # update for the min/max nav attributes in raw_nav
        for k, v in fqpr_instance.multibeam.raw_nav.attrs.items():
            newattrs[k] = v
    except AttributeError:
        print('Unable to read from Navigation')

    if include_mode:
        translated_mode = [mode_translator[a] for a in fqpr_instance.return_unique_mode()]
        newattrs['mode'] = str(translated_mode)
    return newattrs


def write_all_attributes_to_excel(list_dir_paths: list, output_excel: str):
    """
    Using get_attributes_from_zarr_stores, write an excel document, where each row contains the attributes from
    each provided fqpr_generation made fqpr zarr store.

    Parameters
    ----------
    list_dir_paths
        list of strings for paths to each converted folder containing the zarr folders
    output_excel
        path to the newly created excel file
    """

    attrs = get_attributes_from_zarr_stores(list_dir_paths)
    headers = list(attrs[0].keys())

    wb = openpyxl.Workbook()
    for name in wb.get_sheet_names():
        if name == 'Sheet':
            temp = wb.get_sheet_by_name('Sheet')
            wb.remove_sheet(temp)
    ws = wb.create_sheet('Kluster Attributes')

    for cnt, h in enumerate(headers):
        ws[chr(ord('@') + cnt + 1) + '1'] = h
    for row_indx, att in enumerate(attrs):
        for cnt, key in enumerate(att):
            ws[chr(ord('@') + cnt + 1) + str(row_indx + 2)] = str(att[key])
    wb.save(output_excel)
